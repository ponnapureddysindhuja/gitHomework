import csv
from openpyxl import Workbook
from openpyxl.styles import Font
import os

csv_file = 'test_cases.csv'
xlsx_file = 'test_cases.xlsx'

if not os.path.exists(csv_file):
    print(f'CSV not found: {csv_file}')
    raise SystemExit(1)

wb = Workbook()
ws = wb.active

with open(csv_file, newline='', encoding='utf-8') as f:
    reader = csv.reader(f)
    rows = list(reader)

# Write rows
for r_idx, row in enumerate(rows, start=1):
    for c_idx, val in enumerate(row, start=1):
        ws.cell(row=r_idx, column=c_idx, value=val)

# Bold header
if rows:
    for c in range(1, len(rows[0]) + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)

# Auto-fit column widths (approx)
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if cell.value:
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column].width = adjusted_width

wb.save(xlsx_file)
print('Saved', xlsx_file)
